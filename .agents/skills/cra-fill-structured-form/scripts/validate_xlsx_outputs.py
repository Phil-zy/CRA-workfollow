from __future__ import annotations

import argparse
import json

from openpyxl import load_workbook

from common import (
    ALLOWED_STATUSES,
    CHECKLIST_HEADERS,
    checklist_row,
    display_value,
    ensure_within,
    inspect_xlsx,
    load_facts,
    load_json,
    normalize_text,
    resolved,
    sha256_file,
    validate_config,
)
from fill_xlsx import process_fields, workbook_cell


def main() -> None:
    parser = argparse.ArgumentParser(description="验证 CRA Excel 填写结果和 Excel 核对清单")
    parser.add_argument("--target-input", required=True)
    parser.add_argument("--output-xlsx", required=True)
    parser.add_argument("--checklist", required=True)
    parser.add_argument("--config", required=True)
    parser.add_argument("--audit-json", required=True)
    parser.add_argument("--authorized-workspace", required=True)
    args = parser.parse_args()

    workspace = resolved(args.authorized_workspace)
    target = resolved(args.target_input)
    output = resolved(args.output_xlsx)
    checklist = resolved(args.checklist)
    config_path = resolved(args.config)
    audit_path = resolved(args.audit_json)
    ensure_within(workspace, [target, output, checklist, config_path, audit_path])
    if output == target:
        raise ValueError("输出 Excel 与输入路径相同")

    config = load_json(config_path)
    validate_config(config, "enabled")
    if config["template"]["format"] != "xlsx":
        raise RuntimeError("模板配置格式不是 xlsx")
    audit = load_json(audit_path)
    facts_path = resolved(audit.get("facts_input", ""))
    recorded_config_path = resolved(audit.get("config", ""))
    ensure_within(workspace, [facts_path, recorded_config_path])
    checks = (
        (recorded_config_path == config_path, "模板配置路径与审计记录不一致"),
        (resolved(audit.get("target_input", "")) == target, "目标输入路径与审计记录不一致"),
        (resolved(audit.get("output_form", "")) == output, "输出 Excel 路径与审计记录不一致"),
        (resolved(audit.get("output_checklist", "")) == checklist, "核对清单路径与审计记录不一致"),
        (audit.get("output_format") == "xlsx", "审计记录中的输出格式不是 xlsx"),
        (config["config_id"] == audit.get("config_id") and config["version"] == audit.get("config_version"), "模板配置身份或版本与审计记录不一致"),
        (sha256_file(config_path) == audit.get("config_sha256"), "模板配置哈希与审计记录不一致"),
        (sha256_file(facts_path) == audit.get("facts_input_sha256"), "事实输入哈希发生变化"),
        (sha256_file(target) == audit.get("target_input_sha256"), "输入 Excel 哈希发生变化"),
        (sha256_file(output) == audit.get("output_form_sha256"), "输出 Excel 哈希与审计记录不一致"),
        (sha256_file(checklist) == audit.get("output_checklist_sha256"), "核对清单哈希与审计记录不一致"),
    )
    for condition, message in checks:
        if not condition:
            raise RuntimeError(message)
    if not str(audit.get("authorization_id", "")).strip():
        raise RuntimeError("审计记录缺少明确的授权确认 ID")

    source_inspection = inspect_xlsx(target)
    output_inspection = inspect_xlsx(output)
    if source_inspection["geometry"] != output_inspection["geometry"]:
        raise RuntimeError("Excel 工作表、合并区域、公式、样式或打印设置发生非预期变化")

    source_workbook = load_workbook(target, data_only=False, read_only=False, keep_links=True)
    output_workbook = load_workbook(output, data_only=False, read_only=False, keep_links=True)
    try:
        audit_rows = audit.get("fields", [])
        recalculated_rows = process_fields(source_workbook, config, load_facts(facts_path), write=False)
        statuses = [row.get("status") for row in audit_rows]
        if len(statuses) != len(config["fields"]) or any(status not in ALLOWED_STATUSES for status in statuses):
            raise RuntimeError("字段状态数量或取值不符合配置")
        stable_keys = ("field_id", "field_name", "target_location", "status", "source_value", "target_original", "source_file", "source_location")
        allowed_changes = {
            (field["target"]["sheet"], field["target"]["cell"])
            for field, audit_row in zip(config["fields"], audit_rows, strict=True)
            if audit_row.get("status") == "已填写"
        }
        for field, audit_row, recalculated_row in zip(config["fields"], audit_rows, recalculated_rows, strict=True):
            if any(audit_row.get(key, "") != recalculated_row.get(key, "") for key in stable_keys):
                raise RuntimeError(f"审计字段状态或来源与输入事实不一致: {field['label']}")
            before = workbook_cell(source_workbook, field)
            after = workbook_cell(output_workbook, field)
            before_value = display_value(before.value)
            after_value = display_value(after.value)
            if audit_row["status"] == "已填写":
                if normalize_text(after_value) != normalize_text(audit_row.get("source_value", "")):
                    raise RuntimeError(f"Excel 实际写入值与审计记录不一致: {field['label']}")
            elif before_value != after_value:
                raise RuntimeError(f"未标记为已填写的 Excel 字段发生变化: {field['label']}")
        for source_sheet, output_sheet in zip(source_workbook.worksheets, output_workbook.worksheets, strict=True):
            for row in range(1, max(source_sheet.max_row, output_sheet.max_row) + 1):
                for column in range(1, max(source_sheet.max_column, output_sheet.max_column) + 1):
                    source_cell = source_sheet.cell(row=row, column=column)
                    output_cell = output_sheet.cell(row=row, column=column)
                    if (source_sheet.title, source_cell.coordinate) in allowed_changes:
                        continue
                    if display_value(source_cell.value) != display_value(output_cell.value):
                        raise RuntimeError(f"未配置的 Excel 单元格发生变化: {source_sheet.title}!{source_cell.coordinate}")
    finally:
        source_workbook.close()
        output_workbook.close()

    workbook = load_workbook(checklist, data_only=False, read_only=True)
    try:
        if workbook.sheetnames != ["字段核对", "运行记录"]:
            raise RuntimeError("核对清单工作表结构不正确")
        review_sheet = workbook["字段核对"]
        if review_sheet.max_row != len(config["fields"]) + 1 or review_sheet.max_column != 9:
            raise RuntimeError("核对清单字段行数或列数不正确")
        actual_headers = [str(review_sheet.cell(row=1, column=index).value or "") for index in range(1, 10)]
        if actual_headers != CHECKLIST_HEADERS:
            raise RuntimeError("核对清单表头不正确")
        for row_number, audit_row in enumerate(audit_rows, start=2):
            actual = [str(review_sheet.cell(row=row_number, column=index).value or "") for index in range(1, 10)]
            if actual != checklist_row(audit_row):
                raise RuntimeError(f"核对清单第 {row_number} 行内容与审计记录不一致")
        metadata = workbook["运行记录"]
        metadata_values = {str(metadata.cell(row=row, column=1).value or ""): str(metadata.cell(row=row, column=2).value or "") for row in range(2, metadata.max_row + 1)}
        required_metadata = {
            "目标输入": audit["target_input"],
            "目标输入 SHA-256": audit["target_input_sha256"],
            "事实输入": audit["facts_input"],
            "事实输入 SHA-256": audit["facts_input_sha256"],
            "模板配置": audit["config"],
            "模板配置 SHA-256": audit["config_sha256"],
            "表单输出格式": "xlsx",
            "表单输出": audit["output_form"],
            "表单输出 SHA-256": audit["output_form_sha256"],
            "核对清单输出": audit["output_checklist"],
        }
        if any(metadata_values.get(key) != str(value) for key, value in required_metadata.items()):
            raise RuntimeError("核对清单运行记录与审计记录不一致")
    finally:
        workbook.close()

    print(json.dumps({
        "status": "passed",
        "input_unchanged": True,
        "xlsx_structure_preserved": True,
        "formulas_preserved": True,
        "styles_preserved": True,
        "print_settings_preserved": True,
        "status_rows_valid": True,
        "checklist_content_valid": True,
        "xlsx_content_matches_audit": True,
        "visual_render_required": True,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
