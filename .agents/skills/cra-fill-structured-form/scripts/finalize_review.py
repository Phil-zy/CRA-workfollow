from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from common import (
    CHECKLIST_HEADERS,
    checklist_row,
    ensure_output_isolated,
    ensure_within,
    load_json,
    publish_new_file,
    resolved,
    sha256_file,
    temporary_output_path,
    unique_output_path,
    write_json_new,
)


def required_text(value: str, label: str) -> str:
    normalized = str(value or "").strip()
    if not normalized:
        raise ValueError(f"{label}不能为空")
    return normalized


def main() -> None:
    parser = argparse.ArgumentParser(description="把 CRA 人工核对结果记录为不可覆盖的核对清单和验收审计")
    parser.add_argument("--audit-json", required=True)
    parser.add_argument("--checklist", required=True)
    parser.add_argument("--authorized-workspace", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--review-json", required=True)
    parser.add_argument("--reviewed-by", required=True)
    parser.add_argument("--review-id", required=True)
    parser.add_argument("--decision", required=True)
    args = parser.parse_args()

    reviewed_by = required_text(args.reviewed_by, "核对人")
    review_id = required_text(args.review_id, "人工核对 ID")
    decision = required_text(args.decision, "人工核对结果")
    workspace = resolved(args.authorized_workspace)
    audit_path = resolved(args.audit_json)
    checklist_path = resolved(args.checklist)
    output_dir = resolved(args.output_dir)
    review_json_path = resolved(args.review_json)
    ensure_within(workspace, [audit_path, checklist_path, output_dir, review_json_path])
    if review_json_path.exists():
        raise FileExistsError(f"人工核对审计已存在，不得覆盖: {review_json_path}")

    audit = load_json(audit_path)
    source_paths = [resolved(audit[key]) for key in ("target_input", "facts_input", "config")]
    output_form = resolved(audit.get("output_form") or audit.get("output_docx", ""))
    ensure_within(workspace, [*source_paths, output_form])
    ensure_output_isolated(output_dir, source_paths)
    ensure_output_isolated(review_json_path.parent, source_paths)
    if output_dir != checklist_path.parent or output_dir != output_form.parent:
        raise ValueError("人工核对记录必须与本次双输出保存到同一独立输出目录")
    if resolved(audit.get("output_checklist", "")) != checklist_path:
        raise ValueError("核对清单路径与执行审计不一致")
    if sha256_file(checklist_path) != audit.get("output_checklist_sha256"):
        raise ValueError("核对清单哈希与执行审计不一致")
    if sha256_file(output_form) != (audit.get("output_form_sha256") or audit.get("output_docx_sha256")):
        raise ValueError("表单输出哈希与执行审计不一致")
    for input_path, hash_key in zip(
        source_paths,
        ("target_input_sha256", "facts_input_sha256", "config_sha256"),
        strict=True,
    ):
        if sha256_file(input_path) != audit.get(hash_key):
            raise ValueError(f"输入或配置哈希与执行审计不一致: {hash_key}")

    rows = audit.get("fields", [])
    if not rows:
        raise ValueError("执行审计缺少字段记录")
    reviewed_at = datetime.now(timezone.utc).isoformat()
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    target_stem = resolved(audit["target_input"]).stem
    reviewed_checklist = unique_output_path(output_dir, f"{target_stem}_CRA核对记录_{timestamp}.xlsx")
    workbook = load_workbook(checklist_path, data_only=False, read_only=False)
    temporary = temporary_output_path(reviewed_checklist)
    output_published = False
    try:
        if workbook.sheetnames != ["字段核对", "运行记录"]:
            raise ValueError("核对清单工作表结构不正确")
        review_sheet = workbook["字段核对"]
        headers = [str(review_sheet.cell(row=1, column=index).value or "") for index in range(1, 10)]
        if headers != CHECKLIST_HEADERS or review_sheet.max_row != len(rows) + 1:
            raise ValueError("核对清单表头或字段行数不正确")
        field_decisions: list[dict[str, str]] = []
        for row_number, audit_row in enumerate(rows, start=2):
            actual = [str(review_sheet.cell(row=row_number, column=index).value or "") for index in range(1, 10)]
            if actual != checklist_row(audit_row):
                raise ValueError(f"核对清单第 {row_number} 行与执行审计不一致")
            review_sheet.cell(row=row_number, column=9, value=decision)
            field_decisions.append(
                {
                    "field_id": str(audit_row.get("field_id", "")),
                    "field_name": str(audit_row.get("field_name", "")),
                    "status": str(audit_row.get("status", "")),
                    "cra_final_decision": decision,
                }
            )

        metadata = workbook["运行记录"]
        entries = [
            ("人工核对时间", reviewed_at),
            ("核对人", reviewed_by),
            ("人工核对 ID", review_id),
            ("人工核对结果", decision),
            ("原核对清单", str(checklist_path)),
            ("原核对清单 SHA-256", sha256_file(checklist_path)),
            ("已核对清单输出", str(reviewed_checklist)),
            ("人工核对审计", str(review_json_path)),
        ]
        for entry in entries:
            metadata.append(entry)
        for row in metadata.iter_rows(min_row=metadata.max_row - len(entries) + 1):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        workbook.save(temporary)
        reviewed_hash = sha256_file(temporary)
        publish_new_file(temporary, reviewed_checklist)
        output_published = True
        review_record = {
            "schema_version": "1.0",
            "reviewed_at": reviewed_at,
            "reviewed_by": reviewed_by,
            "review_id": review_id,
            "overall_decision": decision,
            "execution_audit": str(audit_path),
            "execution_audit_sha256": sha256_file(audit_path),
            "target_input": audit["target_input"],
            "target_input_sha256": audit["target_input_sha256"],
            "facts_input": audit["facts_input"],
            "facts_input_sha256": audit["facts_input_sha256"],
            "config": audit["config"],
            "config_sha256": audit["config_sha256"],
            "config_id": audit["config_id"],
            "config_version": audit["config_version"],
            "output_form": str(output_form),
            "output_form_sha256": sha256_file(output_form),
            "original_checklist": str(checklist_path),
            "original_checklist_sha256": sha256_file(checklist_path),
            "reviewed_checklist": str(reviewed_checklist),
            "reviewed_checklist_sha256": reviewed_hash,
            "field_decisions": field_decisions,
        }
        write_json_new(review_json_path, review_record)
    except Exception:
        if output_published and reviewed_checklist.exists():
            reviewed_checklist.unlink()
        raise
    finally:
        workbook.close()
        if temporary.exists():
            temporary.unlink()

    print(
        json.dumps(
            {
                "status": "completed",
                "reviewed_checklist": str(reviewed_checklist),
                "review_json": str(review_json_path),
                "field_count": len(rows),
                "decision": decision,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
