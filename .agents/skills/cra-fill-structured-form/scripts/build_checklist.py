from __future__ import annotations

import argparse
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from common import (
    ALLOWED_STATUSES,
    CHECKLIST_HEADERS,
    checklist_row,
    ensure_output_isolated,
    ensure_within,
    load_json,
    publish_new_file,
    resolved,
    sha256_file,
    temporary_output_path,
    unique_output_path,
    update_json_atomic,
)


STATUS_FILLS = {
    "已填写": "E2F0D9",
    "已核对": "DDEBF7",
    "缺失": "FFF2CC",
    "冲突": "F4CCCC",
    "待确认": "FCE4D6",
    "人工保留": "E7E6E6",
}


def main() -> None:
    parser = argparse.ArgumentParser(description="从内部审计记录生成 CRA 填写核对清单")
    parser.add_argument("--audit-json", required=True)
    parser.add_argument("--authorized-workspace", required=True)
    parser.add_argument("--output-dir", required=True)
    args = parser.parse_args()

    workspace = resolved(args.authorized_workspace)
    audit_path = resolved(args.audit_json)
    output_dir = resolved(args.output_dir)
    ensure_within(workspace, [audit_path, output_dir])
    audit = load_json(audit_path)
    input_paths = [resolved(audit[key]) for key in ("target_input", "facts_input", "config")]
    ensure_within(workspace, input_paths)
    ensure_output_isolated(output_dir, input_paths)
    output_form = audit.get("output_form") or audit.get("output_docx")
    output_form_sha256 = audit.get("output_form_sha256") or audit.get("output_docx_sha256")
    output_format = audit.get("output_format") or "docx"
    if not output_form or not output_form_sha256 or output_format not in {"docx", "xlsx"}:
        raise ValueError("审计记录缺少有效的表单输出信息")
    if output_dir != resolved(output_form).parent:
        raise ValueError("核对清单必须与本次表单输出保存到同一独立输出目录")
    if not str(audit.get("authorization_id", "")).strip():
        raise ValueError("审计记录缺少明确的授权确认 ID")
    rows = audit.get("fields", [])
    if not rows or any(row.get("status") not in ALLOWED_STATUSES for row in rows):
        raise ValueError("审计记录缺少字段或包含非法状态")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "字段核对"
    sheet.append(CHECKLIST_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for item in rows:
        sheet.append(checklist_row(item))
        status_cell = sheet.cell(row=sheet.max_row, column=3)
        status_cell.fill = PatternFill("solid", fgColor=STATUS_FILLS[item["status"]])
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [28, 18, 12, 36, 36, 32, 34, 34, 24]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    metadata = workbook.create_sheet("运行记录")
    metadata.append(["项目", "值"])
    for cell in metadata[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    target_stem = resolved(audit["target_input"]).stem
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    output = unique_output_path(output_dir, f"{target_stem}_填写核对清单_{timestamp}.xlsx")
    entries = [
        ("执行时间", audit.get("executed_at", "")),
        ("授权确认 ID", audit.get("authorization_id", "")),
        ("联网", "否" if not audit.get("network_used") else "是"),
        ("外部服务", "否" if not audit.get("external_services_used") else "是"),
        ("目标输入", audit.get("target_input", "")),
        ("目标输入 SHA-256", audit.get("target_input_sha256", "")),
        ("事实输入", audit.get("facts_input", "")),
        ("事实输入 SHA-256", audit.get("facts_input_sha256", "")),
        ("模板配置", audit.get("config", "")),
        ("模板配置 SHA-256", audit.get("config_sha256", "")),
        ("模板配置版本", f"{audit.get('config_id', '')} v{audit.get('config_version', '')}"),
        ("表单输出格式", output_format),
        ("表单输出", output_form),
        ("表单输出 SHA-256", output_form_sha256),
        ("核对清单输出", str(output)),
        ("核对清单完整性", "SHA-256 记录于任务审计 JSON"),
    ]
    for entry in entries:
        metadata.append(entry)
    metadata.column_dimensions["A"].width = 24
    metadata.column_dimensions["B"].width = 90
    for row in metadata.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    temporary = temporary_output_path(output)
    output_published = False
    try:
        workbook.save(temporary)
        checklist_hash = sha256_file(temporary)
        publish_new_file(temporary, output)
        output_published = True
        audit["output_checklist"] = str(output)
        audit["output_checklist_sha256"] = checklist_hash
        update_json_atomic(audit_path, audit)
    except Exception:
        if output_published and output.exists():
            output.unlink()
        raise
    finally:
        if temporary.exists():
            temporary.unlink()
    print(json.dumps({"status": "completed", "checklist": str(output), "field_count": len(rows)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
