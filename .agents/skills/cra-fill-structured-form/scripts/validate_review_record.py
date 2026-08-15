from __future__ import annotations

import argparse
import json

from openpyxl import load_workbook

from common import CHECKLIST_HEADERS, checklist_row, ensure_within, load_json, resolved, sha256_file


def main() -> None:
    parser = argparse.ArgumentParser(description="验证 CRA 人工核对清单与验收审计的完整关联")
    parser.add_argument("--review-json", required=True)
    parser.add_argument("--authorized-workspace", required=True)
    args = parser.parse_args()

    workspace = resolved(args.authorized_workspace)
    review_path = resolved(args.review_json)
    ensure_within(workspace, [review_path])
    review = load_json(review_path)
    required_text = ("reviewed_at", "reviewed_by", "review_id", "overall_decision")
    if any(not str(review.get(key, "")).strip() for key in required_text):
        raise RuntimeError("人工核对审计缺少确认人、时间、ID 或结果")

    path_keys = (
        "execution_audit",
        "target_input",
        "facts_input",
        "config",
        "output_form",
        "original_checklist",
        "reviewed_checklist",
    )
    paths = {key: resolved(review.get(key, "")) for key in path_keys}
    ensure_within(workspace, paths.values())
    hash_checks = (
        ("execution_audit", "execution_audit_sha256"),
        ("target_input", "target_input_sha256"),
        ("facts_input", "facts_input_sha256"),
        ("config", "config_sha256"),
        ("output_form", "output_form_sha256"),
        ("original_checklist", "original_checklist_sha256"),
        ("reviewed_checklist", "reviewed_checklist_sha256"),
    )
    for path_key, hash_key in hash_checks:
        if sha256_file(paths[path_key]) != review.get(hash_key):
            raise RuntimeError(f"人工核对审计中的文件哈希不一致: {path_key}")

    audit = load_json(paths["execution_audit"])
    audit_form = resolved(audit.get("output_form") or audit.get("output_docx", ""))
    checks = (
        (resolved(audit.get("target_input", "")) == paths["target_input"], "目标输入路径不一致"),
        (resolved(audit.get("facts_input", "")) == paths["facts_input"], "事实输入路径不一致"),
        (resolved(audit.get("config", "")) == paths["config"], "模板配置路径不一致"),
        (audit_form == paths["output_form"], "表单输出路径不一致"),
        (resolved(audit.get("output_checklist", "")) == paths["original_checklist"], "原核对清单路径不一致"),
        (audit.get("target_input_sha256") == review.get("target_input_sha256"), "目标输入哈希记录不一致"),
        (audit.get("facts_input_sha256") == review.get("facts_input_sha256"), "事实输入哈希记录不一致"),
        (audit.get("config_sha256") == review.get("config_sha256"), "模板配置哈希记录不一致"),
        (
            (audit.get("output_form_sha256") or audit.get("output_docx_sha256")) == review.get("output_form_sha256"),
            "表单输出哈希记录不一致",
        ),
        (audit.get("output_checklist_sha256") == review.get("original_checklist_sha256"), "原核对清单哈希记录不一致"),
        (audit.get("config_id") == review.get("config_id"), "配置 ID 不一致"),
        (audit.get("config_version") == review.get("config_version"), "配置版本不一致"),
    )
    for condition, message in checks:
        if not condition:
            raise RuntimeError(message)

    audit_rows = audit.get("fields", [])
    decisions = review.get("field_decisions", [])
    if len(audit_rows) != len(decisions) or not audit_rows:
        raise RuntimeError("人工核对字段数量与执行审计不一致")
    workbook = load_workbook(paths["reviewed_checklist"], data_only=False, read_only=True)
    try:
        if workbook.sheetnames != ["字段核对", "运行记录"]:
            raise RuntimeError("已核对清单工作表结构不正确")
        review_sheet = workbook["字段核对"]
        headers = [str(review_sheet.cell(row=1, column=index).value or "") for index in range(1, 10)]
        if headers != CHECKLIST_HEADERS or review_sheet.max_row != len(audit_rows) + 1:
            raise RuntimeError("已核对清单表头或字段行数不正确")
        for row_number, (audit_row, field_decision) in enumerate(zip(audit_rows, decisions, strict=True), start=2):
            if any(
                str(field_decision.get(key, "")) != str(audit_row.get(audit_key, ""))
                for key, audit_key in (("field_id", "field_id"), ("field_name", "field_name"), ("status", "status"))
            ):
                raise RuntimeError(f"人工核对字段身份不一致: 第 {row_number} 行")
            expected = checklist_row(audit_row)[:8] + [str(field_decision.get("cra_final_decision", ""))]
            actual = [str(review_sheet.cell(row=row_number, column=index).value or "") for index in range(1, 10)]
            if actual != expected or actual[8] != str(review["overall_decision"]):
                raise RuntimeError(f"人工核对决定不一致: 第 {row_number} 行")

        metadata = workbook["运行记录"]
        metadata_values = {
            str(metadata.cell(row=row, column=1).value or ""): str(metadata.cell(row=row, column=2).value or "")
            for row in range(2, metadata.max_row + 1)
        }
        required_metadata = {
            "人工核对时间": review["reviewed_at"],
            "核对人": review["reviewed_by"],
            "人工核对 ID": review["review_id"],
            "人工核对结果": review["overall_decision"],
            "原核对清单": review["original_checklist"],
            "原核对清单 SHA-256": review["original_checklist_sha256"],
            "已核对清单输出": review["reviewed_checklist"],
            "人工核对审计": str(review_path),
        }
        if any(metadata_values.get(key) != str(value) for key, value in required_metadata.items()):
            raise RuntimeError("已核对清单运行记录与人工核对审计不一致")
    finally:
        workbook.close()

    print(
        json.dumps(
            {
                "status": "passed",
                "confirmation_traceable": True,
                "input_output_hashes_valid": True,
                "field_decisions_valid": True,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
