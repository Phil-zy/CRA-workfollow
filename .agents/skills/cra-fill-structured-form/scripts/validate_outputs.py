from __future__ import annotations

import argparse
import json

from docx import Document
from openpyxl import load_workbook

from common import (
    ALLOWED_STATUSES,
    CHECKLIST_HEADERS,
    checklist_row,
    ensure_within,
    inspect_docx,
    load_facts,
    load_json,
    normalize_text,
    resolved,
    selected_options,
    sha256_file,
    target_cell,
    validate_config,
)
from fill_docx import process_fields


def validate_written_field(field: dict, audit_row: dict, source_document: Document, output_document: Document) -> None:
    before = target_cell(source_document, field["target"]).text
    after = target_cell(output_document, field["target"]).text
    status = audit_row["status"]
    if before != audit_row.get("target_original", ""):
        raise RuntimeError(f"审计中的目标原值与输入 Word 不一致: {field['label']}")
    if status != "已填写":
        if before != after:
            raise RuntimeError(f"未标记为已填写的字段发生变化: {field['label']}")
        return

    mapping = field["mapping"]
    source_value = str(audit_row.get("source_value", ""))
    if mapping["type"] == "direct":
        if mapping.get("mode") == "inline_after_anchor":
            expected = mapping["anchor"] + source_value
            if normalize_text(expected) != normalize_text(after):
                raise RuntimeError(f"Word 实际写入值与审计记录不一致: {field['label']}")
        elif normalize_text(after) != normalize_text(source_value):
            raise RuntimeError(f"Word 实际写入值与审计记录不一致: {field['label']}")
    elif mapping["type"] == "choice":
        expected = mapping["value_to_option"].get(normalize_text(source_value))
        selected = selected_options(after, mapping["value_to_option"].values(), mapping["checked"])
        if selected != [expected]:
            raise RuntimeError(f"Word 实际勾选值与审计记录不一致: {field['label']}")
    else:
        raise RuntimeError(f"字段状态与映射类型不一致: {field['label']}")


def main() -> None:
    parser = argparse.ArgumentParser(description="验证 CRA Word 填写结果和 Excel 核对清单")
    parser.add_argument("--target-input", required=True)
    parser.add_argument("--output-docx", required=True)
    parser.add_argument("--checklist", required=True)
    parser.add_argument("--config", required=True)
    parser.add_argument("--audit-json", required=True)
    parser.add_argument("--authorized-workspace", required=True)
    args = parser.parse_args()

    workspace = resolved(args.authorized_workspace)
    target = resolved(args.target_input)
    output = resolved(args.output_docx)
    checklist = resolved(args.checklist)
    config_path = resolved(args.config)
    audit_path = resolved(args.audit_json)
    ensure_within(workspace, [target, output, checklist, config_path, audit_path])
    if output == target:
        raise ValueError("输出 Word 与输入路径相同")

    config = load_json(config_path)
    validate_config(config, "enabled")
    audit = load_json(audit_path)
    facts_path = resolved(audit.get("facts_input", ""))
    recorded_config_path = resolved(audit.get("config", ""))
    ensure_within(workspace, [facts_path, recorded_config_path])
    if recorded_config_path != config_path:
        raise RuntimeError("模板配置路径与审计记录不一致")
    if resolved(audit.get("target_input", "")) != target:
        raise RuntimeError("目标输入路径与审计记录不一致")
    if not str(audit.get("authorization_id", "")).strip():
        raise RuntimeError("审计记录缺少明确的授权确认 ID")
    if config["config_id"] != audit.get("config_id") or config["version"] != audit.get("config_version"):
        raise RuntimeError("模板配置身份或版本与审计记录不一致")
    if sha256_file(config_path) != audit.get("config_sha256"):
        raise RuntimeError("模板配置哈希与审计记录不一致")
    if sha256_file(facts_path) != audit.get("facts_input_sha256"):
        raise RuntimeError("事实输入哈希发生变化")
    if sha256_file(target) != audit["target_input_sha256"]:
        raise RuntimeError("输入 Word 哈希发生变化")
    if audit.get("output_format") != "docx":
        raise RuntimeError("审计记录中的输出格式不是 docx")
    if resolved(audit.get("output_docx", "")) != output:
        raise RuntimeError("输出 Word 路径与审计记录不一致")
    if resolved(audit.get("output_form", "")) != output or audit.get("output_form_sha256") != audit.get("output_docx_sha256"):
        raise RuntimeError("通用表单输出记录与 Word 输出记录不一致")
    if sha256_file(output) != audit["output_docx_sha256"]:
        raise RuntimeError("输出 Word 哈希与审计记录不一致")
    if resolved(audit.get("output_checklist", "")) != checklist:
        raise RuntimeError("核对清单路径与审计记录不一致")
    if sha256_file(checklist) != audit.get("output_checklist_sha256"):
        raise RuntimeError("核对清单哈希与审计记录不一致")

    source_inspection = inspect_docx(target)
    output_inspection = inspect_docx(output)
    if source_inspection["geometry"] != output_inspection["geometry"]:
        raise RuntimeError("Word 段落、表格或合并单元格结构发生变化")

    source_document = Document(target)
    output_document = Document(output)
    audit_rows = audit.get("fields", [])
    recalculated_rows = process_fields(source_document, config, load_facts(facts_path), write=False)
    statuses = [row.get("status") for row in audit_rows]
    if len(statuses) != len(config["fields"]) or any(status not in ALLOWED_STATUSES for status in statuses):
        raise RuntimeError("字段状态数量或取值不符合配置")
    stable_keys = (
        "field_id",
        "field_name",
        "target_location",
        "status",
        "source_value",
        "target_original",
        "source_file",
        "source_location",
    )
    for field, audit_row, recalculated_row in zip(config["fields"], audit_rows, recalculated_rows, strict=True):
        if audit_row.get("field_id") != field["id"] or audit_row.get("field_name") != field["label"]:
            raise RuntimeError("审计字段顺序或身份与模板配置不一致")
        if any(audit_row.get(key, "") != recalculated_row.get(key, "") for key in stable_keys):
            raise RuntimeError(f"审计字段状态或来源与输入事实不一致: {field['label']}")
        validate_written_field(field, audit_row, source_document, output_document)

    workbook = load_workbook(checklist, data_only=False, read_only=True)
    if workbook.sheetnames != ["字段核对", "运行记录"]:
        raise RuntimeError("核对清单工作表结构不正确")
    review_sheet = workbook["字段核对"]
    if review_sheet.max_row != len(config["fields"]) + 1 or review_sheet.max_column != 9:
        raise RuntimeError("核对清单字段行数或列数不正确")
    actual_headers = [str(review_sheet.cell(row=1, column=index).value or "") for index in range(1, 10)]
    if actual_headers != CHECKLIST_HEADERS:
        raise RuntimeError("核对清单表头不正确")
    for row_number, audit_row in enumerate(audit_rows, start=2):
        actual = [str(review_sheet.cell(row=row_number, column=index).value or "") for index in range(1, 10)]
        if actual != checklist_row(audit_row):
            raise RuntimeError(f"核对清单第 {row_number} 行内容与审计记录不一致")
    metadata = workbook["运行记录"]
    metadata_values = {
        str(metadata.cell(row=row, column=1).value or ""): str(metadata.cell(row=row, column=2).value or "")
        for row in range(2, metadata.max_row + 1)
    }
    required_metadata = {
        "目标输入": audit["target_input"],
        "目标输入 SHA-256": audit["target_input_sha256"],
        "事实输入": audit["facts_input"],
        "事实输入 SHA-256": audit["facts_input_sha256"],
        "模板配置": audit["config"],
        "模板配置 SHA-256": audit["config_sha256"],
        "表单输出格式": "docx",
        "表单输出": audit["output_docx"],
        "表单输出 SHA-256": audit["output_docx_sha256"],
        "核对清单输出": audit["output_checklist"],
    }
    if any(metadata_values.get(key) != str(value) for key, value in required_metadata.items()):
        raise RuntimeError("核对清单运行记录与审计记录不一致")
    workbook.close()

    result = {
        "status": "passed",
        "input_unchanged": True,
        "docx_geometry_preserved": True,
        "manual_regions_preserved": True,
        "status_rows_valid": True,
        "checklist_structure_valid": True,
        "checklist_content_valid": True,
        "word_content_matches_audit": True,
        "visual_render_required": True,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
