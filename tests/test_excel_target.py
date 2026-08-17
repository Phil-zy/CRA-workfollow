from __future__ import annotations

import json
import hashlib
import os
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / ".agents" / "skills" / "cra-fill-structured-form" / "scripts"))
import build_checklist as checklist_module
import common as common_module
import fill_xlsx as fill_xlsx_module


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / ".agents" / "skills" / "cra-fill-structured-form" / "scripts"


def run_script(name: str, *args: object, expect_success: bool = True) -> subprocess.CompletedProcess[str]:
    environment = os.environ.copy()
    environment["PYTHONIOENCODING"] = "utf-8"
    process = subprocess.run(
        [sys.executable, str(SCRIPTS / name), *(str(arg) for arg in args)],
        cwd=ROOT,
        text=True,
        encoding="utf-8",
        capture_output=True,
        check=False,
        env=environment,
    )
    if expect_success and process.returncode != 0:
        raise AssertionError(f"{name} failed:\nSTDOUT:\n{process.stdout}\nSTDERR:\n{process.stderr}")
    if not expect_success and process.returncode == 0:
        raise AssertionError(f"{name} unexpectedly succeeded:\n{process.stdout}")
    return process


class ExcelTargetTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.workspace = Path(self.temp.name)
        self.target = self.workspace / "site-startup-form.xlsx"
        self.config_library = self.workspace / "template-configs"
        self.facts = self.workspace / "facts.xlsx"
        self.output_dir = self.workspace / "output"
        self.audit_dir = self.workspace / "audit"
        self.audit = self.audit_dir / "audit.json"
        self._write_template(self.target)
        self._write_facts(self.facts)

    def tearDown(self) -> None:
        self.temp.cleanup()

    @staticmethod
    def _write_template(path: Path, *, protected: bool = False) -> None:
        workbook = Workbook()
        form = workbook.active
        form.title = "申请信息"
        form.merge_cells("A1:D1")
        form["A1"] = "临床试验项目启动信息表"
        form["A1"].font = Font(bold=True, size=16)
        form["A1"].fill = PatternFill("solid", fgColor="1F4E78")
        form["A3"] = "项目名称"
        form["A4"] = "方案编号"
        form["A5"] = "方案版本号"
        form["B5"].number_format = "@"
        form["A6"] = "经费来源"
        form["B5"] = "2.0"
        form["B6"] = "政府立项"
        form["D3"] = "已填写字段数"
        form["D4"] = "=COUNTA(B3:B6)"
        form.page_setup.orientation = "landscape"
        form.print_area = "A1:D8"

        site = workbook.create_sheet("中心信息")
        site["A2"] = "研究科室"
        site["A3"] = "主要研究者"
        site["A4"] = "主要研究者签名"
        site.merge_cells("B2:C2")
        site.page_setup.orientation = "portrait"
        site.print_area = "A1:C6"
        if protected:
            site.protection.sheet = True
        workbook.save(path)

    @staticmethod
    def _write_facts(path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "项目信息"
        sheet.append(["字段", "内容"])
        sheet.append(["项目名称", "测试器械临床试验"])
        sheet.append(["方案编号", "DEV-2026-001"])
        sheet.append(["方案版本号", "2.0"])
        sheet.append(["经费来源", "医药公司"])
        sheet.append(["研究科室", "皮肤科"])
        sheet.append(["主要研究者", "测试研究者"])
        workbook.save(path)

    @staticmethod
    def _sha256(path: Path) -> str:
        return hashlib.sha256(path.read_bytes()).hexdigest()

    def _write_config(self) -> Path:
        inspection = json.loads(self._inspect().stdout)["template"]
        config = {
            "schema_version": "1.0",
            "config_id": "site-startup-form",
            "version": 1,
            "status": "enabled",
            "form_type": "site-startup-form",
            "template": inspection,
            "fields": [
                {
                    "id": "project_name",
                    "label": "项目名称",
                    "fact_key": "项目名称",
                    "target": {"sheet": "申请信息", "cell": "B3", "expected_original": ""},
                    "mapping": {"type": "direct", "mode": "cell"},
                    "mapping_review": "approved",
                },
                {
                    "id": "protocol_number",
                    "label": "方案编号",
                    "fact_key": "方案编号",
                    "target": {"sheet": "申请信息", "cell": "B4", "expected_original": ""},
                    "mapping": {"type": "direct", "mode": "cell"},
                    "mapping_review": "approved",
                },
                {
                    "id": "protocol_version",
                    "label": "方案版本号",
                    "fact_key": "方案版本号",
                    "target": {"sheet": "申请信息", "cell": "B5", "expected_original": "2.0"},
                    "mapping": {"type": "direct", "mode": "cell"},
                    "mapping_review": "approved",
                },
                {
                    "id": "funding_source",
                    "label": "经费来源",
                    "fact_key": "经费来源",
                    "target": {"sheet": "申请信息", "cell": "B6", "expected_original": "政府立项"},
                    "mapping": {"type": "direct", "mode": "cell"},
                    "mapping_review": "approved",
                },
                {
                    "id": "formula_guard",
                    "label": "公式保护验证",
                    "fact_key": "方案版本号",
                    "target": {"sheet": "申请信息", "cell": "D4", "expected_original": "=COUNTA(B3:B6)"},
                    "mapping": {"type": "direct", "mode": "cell"},
                    "mapping_review": "approved",
                },
                {
                    "id": "department",
                    "label": "研究科室",
                    "fact_key": "研究科室",
                    "target": {"sheet": "中心信息", "cell": "B2", "expected_original": ""},
                    "mapping": {"type": "direct", "mode": "cell"},
                    "mapping_review": "approved",
                },
                {
                    "id": "principal_investigator",
                    "label": "主要研究者",
                    "fact_key": "主要研究者",
                    "target": {"sheet": "中心信息", "cell": "B3", "expected_original": ""},
                    "mapping": {"type": "direct", "mode": "cell"},
                    "mapping_review": "approved",
                },
                {
                    "id": "acceptance_number",
                    "label": "项目受理号",
                    "fact_key": "项目受理号",
                    "target": {"sheet": "申请信息", "cell": "C6", "expected_original": ""},
                    "mapping": {"type": "direct", "mode": "cell"},
                    "mapping_review": "approved",
                },
                {
                    "id": "pi_signature",
                    "label": "主要研究者签名",
                    "fact_key": None,
                    "target": {"sheet": "中心信息", "cell": "B4", "expected_original": ""},
                    "mapping": {"type": "manual"},
                    "mapping_review": "approved",
                },
            ],
            "approval": {
                "approved_by": "CRA",
                "approved_at": "2026-08-13T00:00:00+00:00",
                "approval_id": "TEST-EXCEL-APPROVAL",
            },
        }
        directory = self.config_library / "enabled"
        directory.mkdir(parents=True, exist_ok=True)
        path = directory / "site-startup-form-v1.json"
        path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
        return path

    def _write_mapping_spec(self) -> Path:
        path = self.workspace / "mapping-spec.json"
        path.write_text(
            json.dumps(
                {
                    "config_id": "site-startup-form",
                    "form_type": "site-startup-form",
                    "fields": [
                        {
                            "id": "project_name",
                            "label": "项目名称",
                            "fact_key": "项目名称",
                            "target": {"sheet": "申请信息", "cell": "B3"},
                            "mapping": {"type": "direct", "mode": "cell"},
                        },
                        {
                            "id": "department",
                            "label": "研究科室",
                            "fact_key": "研究科室",
                            "target": {"sheet": "中心信息", "cell": "B2"},
                            "mapping": {"type": "direct", "mode": "cell"},
                        },
                        {
                            "id": "pi_signature",
                            "label": "主要研究者签名",
                            "fact_key": None,
                            "target": {"sheet": "中心信息", "cell": "B4"},
                            "mapping": {"type": "manual"},
                        },
                    ],
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        return path

    def _inspect(self, target: Path | None = None, *, expect_success: bool = True) -> subprocess.CompletedProcess[str]:
        return run_script(
            "inspect_template.py",
            "--target",
            target or self.target,
            "--authorized-workspace",
            self.workspace,
            "--config-library",
            self.config_library,
            expect_success=expect_success,
        )

    def test_inspects_multisheet_excel_and_rejects_protected_workbook(self) -> None:
        inspection = json.loads(self._inspect().stdout)
        self.assertEqual(inspection["template"]["format"], "xlsx")
        self.assertEqual(inspection["sheet_count"], 2)
        self.assertEqual(inspection["sheet_names"], ["申请信息", "中心信息"])

        protected = self.workspace / "protected.xlsx"
        self._write_template(protected, protected=True)
        rejected = self._inspect(protected, expect_success=False)
        self.assertIn("保护", rejected.stderr)

        revision_protected = self.workspace / "revision-protected.xlsx"
        self._write_template(revision_protected)
        workbook = load_workbook(revision_protected)
        workbook.security.lockRevision = True
        workbook.save(revision_protected)
        workbook.close()
        rejected = self._inspect(revision_protected, expect_success=False)
        self.assertIn("保护", rejected.stderr)

        macro = self.workspace / "macro.xlsx"
        self._write_template(macro)
        with zipfile.ZipFile(macro, "a") as package:
            package.writestr("xl/vbaProject.bin", b"test macro marker")
        self.assertIn("宏", self._inspect(macro, expect_success=False).stderr)

        external = self.workspace / "external.xlsx"
        self._write_template(external)
        with zipfile.ZipFile(external, "a") as package:
            package.writestr("xl/externalLinks/externalLink1.xml", b"test external link marker")
        self.assertIn("外部数据连接", self._inspect(external, expect_success=False).stderr)

        corrupt = self.workspace / "corrupt.xlsx"
        corrupt.write_bytes(b"not-a-valid-xlsx")
        rejected = self._inspect(corrupt, expect_success=False)
        self.assertTrue("加密" in rejected.stderr or "损坏" in rejected.stderr)

    def test_fills_excel_copy_and_preserves_workbook_structure_and_formatting(self) -> None:
        config = self._write_config()
        input_hash = self._sha256(self.target)
        source = load_workbook(self.target, data_only=False)
        source_style = source["申请信息"]["B3"].style_id
        source_formula = source["申请信息"]["D4"].value
        source_merges = {sheet.title: sorted(str(item) for item in sheet.merged_cells.ranges) for sheet in source.worksheets}
        source_print = {sheet.title: (str(sheet.print_area), sheet.page_setup.orientation) for sheet in source.worksheets}
        source.close()

        result = run_script(
            "fill_xlsx.py",
            "--target",
            self.target,
            "--facts",
            self.facts,
            "--config",
            config,
            "--authorized-workspace",
            self.workspace,
            "--output-dir",
            self.output_dir,
            "--audit-json",
            self.audit,
            "--authorization-id",
            "TEST-AUTHORIZATION",
        )
        output = Path(json.loads(result.stdout)["output_xlsx"])
        self.assertNotEqual(output, self.target)
        self.assertEqual(self._sha256(self.target), input_hash)

        workbook = load_workbook(output, data_only=False)
        self.assertEqual(workbook.sheetnames, ["申请信息", "中心信息"])
        self.assertEqual(workbook["申请信息"]["B3"].value, "测试器械临床试验")
        self.assertEqual(workbook["申请信息"]["B4"].value, "DEV-2026-001")
        self.assertEqual(workbook["中心信息"]["B2"].value, "皮肤科")
        self.assertEqual(workbook["中心信息"]["B3"].value, "测试研究者")
        self.assertEqual(workbook["申请信息"]["B5"].value, "2.0")
        self.assertEqual(workbook["申请信息"]["B6"].value, "政府立项")
        self.assertIsNone(workbook["申请信息"]["C6"].value)
        self.assertIsNone(workbook["中心信息"]["B4"].value)
        self.assertEqual(workbook["申请信息"]["B3"].style_id, source_style)
        self.assertEqual(workbook["申请信息"]["D4"].value, source_formula)
        self.assertEqual(
            {sheet.title: sorted(str(item) for item in sheet.merged_cells.ranges) for sheet in workbook.worksheets},
            source_merges,
        )
        self.assertEqual(
            {sheet.title: (str(sheet.print_area), sheet.page_setup.orientation) for sheet in workbook.worksheets},
            source_print,
        )
        workbook.close()

        audit = json.loads(self.audit.read_text(encoding="utf-8"))
        self.assertEqual(
            {row["field_id"]: row["status"] for row in audit["fields"]},
            {
                "project_name": "已填写",
                "protocol_number": "已填写",
                "protocol_version": "已核对",
                "funding_source": "冲突",
                "formula_guard": "待确认",
                "department": "已填写",
                "principal_investigator": "已填写",
                "acceptance_number": "缺失",
                "pi_signature": "人工保留",
            },
        )

        checklist_result = run_script(
            "build_checklist.py",
            "--audit-json",
            self.audit,
            "--authorized-workspace",
            self.workspace,
            "--output-dir",
            self.output_dir,
        )
        checklist = Path(json.loads(checklist_result.stdout)["checklist"])
        validated = run_script(
            "validate_xlsx_outputs.py",
            "--target-input",
            self.target,
            "--output-xlsx",
            output,
            "--checklist",
            checklist,
            "--config",
            config,
            "--audit-json",
            self.audit,
            "--authorized-workspace",
            self.workspace,
        )
        self.assertEqual(json.loads(validated.stdout)["status"], "passed")

        second_audit = self.audit_dir / "second-audit.json"
        second = run_script(
            "fill_xlsx.py",
            "--target",
            self.target,
            "--facts",
            self.facts,
            "--config",
            config,
            "--authorized-workspace",
            self.workspace,
            "--output-dir",
            self.output_dir,
            "--audit-json",
            second_audit,
            "--authorization-id",
            "TEST-AUTHORIZATION",
        )
        second_output = Path(json.loads(second.stdout)["output_xlsx"])
        self.assertNotEqual(second_output, output)
        self.assertEqual(
            json.loads(second_audit.read_text(encoding="utf-8"))["fields"],
            json.loads(self.audit.read_text(encoding="utf-8"))["fields"],
        )
        self.assertEqual(load_workbook(second_output)["申请信息"]["B3"].value, "测试器械临床试验")

        tampered = load_workbook(output)
        tampered["申请信息"]["A3"] = "被篡改的标签"
        tampered.save(output)
        tampered.close()
        audit = json.loads(self.audit.read_text(encoding="utf-8"))
        audit["output_form_sha256"] = self._sha256(output)
        self.audit.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
        rejected = run_script(
            "validate_xlsx_outputs.py",
            "--target-input",
            self.target,
            "--output-xlsx",
            output,
            "--checklist",
            checklist,
            "--config",
            config,
            "--audit-json",
            self.audit,
            "--authorized-workspace",
            self.workspace,
            expect_success=False,
        )
        self.assertIn("未配置的 Excel 单元格发生变化", rejected.stderr)

    def test_checklist_failure_does_not_publish_untracked_final_file(self) -> None:
        config = self._write_config()
        run_script(
            "fill_xlsx.py",
            "--target",
            self.target,
            "--facts",
            self.facts,
            "--config",
            config,
            "--authorized-workspace",
            self.workspace,
            "--output-dir",
            self.output_dir,
            "--audit-json",
            self.audit,
            "--authorization-id",
            "TEST-AUTHORIZATION",
        )
        arguments = [
            "build_checklist.py",
            "--audit-json",
            str(self.audit),
            "--authorized-workspace",
            str(self.workspace),
            "--output-dir",
            str(self.output_dir),
        ]
        with patch.object(sys, "argv", arguments), patch.object(
            checklist_module, "update_json_atomic", side_effect=OSError("simulated audit failure")
        ):
            with self.assertRaises(OSError):
                checklist_module.main()
        self.assertEqual(list(self.output_dir.glob("*填写核对清单*.xlsx")), [])

    def test_audit_failure_does_not_publish_untracked_excel_file(self) -> None:
        config = self._write_config()
        arguments = [
            "fill_xlsx.py",
            "--target",
            str(self.target),
            "--facts",
            str(self.facts),
            "--config",
            str(config),
            "--authorized-workspace",
            str(self.workspace),
            "--output-dir",
            str(self.output_dir),
            "--audit-json",
            str(self.audit),
            "--authorization-id",
            "TEST-AUTHORIZATION",
        ]
        def fail_after_partial_write(payload: dict, handle: object, **kwargs: object) -> None:
            handle.write('{"partial":')
            raise OSError("simulated partial audit failure")

        with patch.object(sys, "argv", arguments), patch.object(
            common_module.json, "dump", side_effect=fail_after_partial_write
        ):
            with self.assertRaises(OSError):
                fill_xlsx_module.main()
        self.assertEqual(list(self.output_dir.glob("*.xlsx")), [])
        self.assertFalse(self.audit.exists())
        self.assertEqual(list(self.audit_dir.iterdir()), [])

    def test_exclusive_publish_never_overwrites_racing_file(self) -> None:
        temporary = self.workspace / "complete.tmp"
        final = self.workspace / "racing.xlsx"
        temporary.write_bytes(b"complete output")
        final.write_bytes(b"existing output")
        with self.assertRaises(FileExistsError):
            common_module.publish_new_file(temporary, final)
        self.assertEqual(final.read_bytes(), b"existing output")
        self.assertEqual(temporary.read_bytes(), b"complete output")

    def test_cra_review_confirmation_creates_a_traceable_immutable_record(self) -> None:
        config = self._write_config()
        fill_result = run_script(
            "fill_xlsx.py",
            "--target",
            self.target,
            "--facts",
            self.facts,
            "--config",
            config,
            "--authorized-workspace",
            self.workspace,
            "--output-dir",
            self.output_dir,
            "--audit-json",
            self.audit,
            "--authorization-id",
            "TEST-AUTHORIZATION",
        )
        output_form = Path(json.loads(fill_result.stdout)["output_xlsx"])
        checklist_result = run_script(
            "build_checklist.py",
            "--audit-json",
            self.audit,
            "--authorized-workspace",
            self.workspace,
            "--output-dir",
            self.output_dir,
        )
        original_checklist = Path(json.loads(checklist_result.stdout)["checklist"])
        original_hash = self._sha256(original_checklist)
        review_json = self.audit_dir / "review.json"

        finalized = run_script(
            "finalize_review.py",
            "--audit-json",
            self.audit,
            "--checklist",
            original_checklist,
            "--authorized-workspace",
            self.workspace,
            "--output-dir",
            self.output_dir,
            "--review-json",
            review_json,
            "--reviewed-by",
            "CRA",
            "--review-id",
            "TEST-REVIEW-001",
            "--decision",
            "核对无误",
        )
        reviewed_checklist = Path(json.loads(finalized.stdout)["reviewed_checklist"])
        self.assertEqual(self._sha256(original_checklist), original_hash)
        workbook = load_workbook(reviewed_checklist, data_only=False, read_only=True)
        try:
            decisions = [workbook["字段核对"].cell(row=row, column=9).value for row in range(2, 11)]
            self.assertEqual(decisions, ["核对无误"] * 9)
            metadata = {
                workbook["运行记录"].cell(row=row, column=1).value: workbook["运行记录"].cell(row=row, column=2).value
                for row in range(2, workbook["运行记录"].max_row + 1)
            }
            self.assertEqual(metadata["人工核对 ID"], "TEST-REVIEW-001")
            self.assertEqual(metadata["人工核对结果"], "核对无误")
        finally:
            workbook.close()

        review = json.loads(review_json.read_text(encoding="utf-8"))
        self.assertEqual(review["output_form"], str(output_form.resolve()))
        self.assertEqual(review["original_checklist_sha256"], original_hash)
        self.assertEqual(review["reviewed_checklist"], str(reviewed_checklist.resolve()))
        self.assertEqual(len(review["field_decisions"]), 9)
        validated = run_script(
            "validate_review_record.py",
            "--review-json",
            review_json,
            "--authorized-workspace",
            self.workspace,
        )
        self.assertEqual(json.loads(validated.stdout)["status"], "passed")

    def test_repository_excel_sample_matches_latest_draft_and_maps_merged_cell(self) -> None:
        target = ROOT / "fixtures" / "form-fill-pilot" / "template" / "模拟项目启动信息表.xlsx"
        inspection = json.loads(
            run_script(
                "inspect_template.py",
                "--target",
                target,
                "--authorized-workspace",
                ROOT,
                "--config-library",
                ROOT / "template-configs" / "cra-fill-structured-form",
            ).stdout
        )
        self.assertTrue(inspection["matching_draft_configs"])
        latest = max((Path(path) for path in inspection["matching_draft_configs"]), key=lambda path: path.name)
        draft = json.loads(latest.read_text(encoding="utf-8"))
        department = next(field for field in draft["fields"] if field["id"] == "department")
        self.assertEqual(department["target"], {"sheet": "中心信息", "cell": "B3", "expected_original": ""})
        workbook = load_workbook(target)
        self.assertIn("B3:C3", {str(item) for item in workbook["中心信息"].merged_cells.ranges})
        workbook.close()

    def test_excel_template_creates_reviewable_mapping_draft_without_filling(self) -> None:
        mapping_spec = self._write_mapping_spec()
        result = run_script(
            "draft_mapping.py",
            "--target",
            self.target,
            "--facts",
            self.facts,
            "--authorized-workspace",
            self.workspace,
            "--config-library",
            self.config_library,
            "--mapping-spec",
            mapping_spec,
        )
        payload = json.loads(result.stdout)
        draft = json.loads(Path(payload["draft"]).read_text(encoding="utf-8"))
        self.assertEqual(payload["next_action"], "停止。向 CRA 展示映射草稿并等待明确批准。")
        self.assertEqual(draft["status"], "draft")
        self.assertEqual(draft["template"]["format"], "xlsx")
        self.assertEqual(draft["fields"][0]["target"], {"sheet": "申请信息", "cell": "B3", "expected_original": ""})
        self.assertFalse(self.output_dir.exists())


if __name__ == "__main__":
    unittest.main()
